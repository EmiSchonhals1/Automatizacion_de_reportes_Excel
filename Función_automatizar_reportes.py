"""ESTE PROGRAMA AUTOMATIZA LOS REPORTES DE LAS VENTAS DE CADA MES. ES OBLIGATORIO QUE EL ARCHIVO DE INGRESO (INPUT) TENGA EL FORMATO Ventas_mes.xlsx PARA EL CORRECTO FUNCIONAMIENTO DEL PROGRAMA.
EL MISMO DEVUELVE COMO SALIDA (OUTPUT) UN ARCHIVO CON EL FORMATO Reporte_mes.xlsx CON SU RESPECTIVO GRÁFICO"""




import pandas as pd #usamos pandas para la lectura y manejo de archivos Excel, creación de tablas pivote y exportación de las mismas a archivos Excel.
from openpyxl import load_workbook #usamos openpyxl para hacer reportes en Excel (usamos load_workbook para leer los archivos Excel con openpyxl)
from openpyxl.chart import BarChart, Reference
#importamos string para utilizar string.ascii_uppercase
import string
#importamos Font para darle formato final al excel del wb
from openpyxl.styles import Font


#creamos la función que nos permitirá automatizar los reportes del Excel
def automatizar_excel(nombre_archivo) :
    """Input Ventas_mes.xlsx / Output reporte_mes.xlsx"""
    
    #leemos el archivo excel (o df, lo que prefiramos) y lo asignamos a una variable
    archivo_excel = pd.read_excel(nombre_archivo) #modificamos el nombre original por esta variable para que se pueda usar ante cualquier archivo y funcione la automatización

    #creamos una tabla pivote en la que ordenaremos el total gastado en cada linea de productos separados por género
    tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    #round(0) es para redondear los decimales
    
    #tomamos como variable el mes y la extensión del archivo para poder hacer el reporte de cada mes que ingrese
    #tomará lo que está luego del _ 
    mes_extension = nombre_archivo.split('_')[1]

    #exportamos la tabla pivote como archivo Excel y le ponemos el nombre que tendrá el archivo Excel (con su terminación .xlsx) y su mes_extension
    tabla_pivote.to_excel(f'Reporte_{mes_extension}', startrow=3, sheet_name='Reporte por género')
    #startrow es el parámetro que controla en que fila (row) del Excel comenzará a mostrarse la tabla_pivote
    #sheet_name es el nombre que le pondremos a la pestaña o ventana de excel


    #------------------------------Creación de reportes de Excel con openpyxl------------------------------
    #leemos el Excel de la tabla pivote con load_workbook (en general le daremos el valor de lectura a una variable wb(WorkBook))
    wb = load_workbook(f'Reporte_{mes_extension}')
    #usamos una variable pestaña y ponemos el nombre de la pestaña o ventana que usaremos del archivo que estamos leyendo (el que pusimos de sheet_name al exportar la tabla pivote a Excel)
    pestaña = wb['Reporte por género']

    #ahora debemos saber que filas y columnas están activas en el excel que estamos utilizando para tener las referencias
    #para eso debemos conocer las filas maximas y minimas y las columnas maximas y minimas que se estan utilizando
    min_columna = wb.active.min_column
    max_columna = wb.active.max_column
    min_fila = wb.active.min_row
    max_fila = wb.active.max_row
    #CON ESTO PODEMOS AUTOMATIZAR CUALQUIER REPORTE, YA QUE AUTOMATICAMENTE EL PROGRAMA TOMARÁ LOS VALORES DE LAS FILAS Y COLUMNAS ACTIVAS, SIN IMPORTAR SI CAMBIAN LUEGO DE TAMAÑO


    #especificamos que la data y categorias tomarán los valores de la pestaña con la que se está trabajando, y los valores mínimos y máximos de las filas y columnas que calculamos antes.

    #queremos que la data tome también los headers o encabezados de la tabla_pivote de excel, por ende le agregamos 1 a la min_columna calculada previamente
    data = Reference(pestaña, min_col=min_columna+1, max_col=max_columna, min_row=min_fila, max_row=max_fila)
    #queremos que las categorias tomen también los valores de las categorias de la tabla_pivote de excel, (por ende le agregamos 1 a la min_fila calculada previamente y también hacemos que la min_col y la max_col tengan el mismo valor de min_columna, ya que solo queremos que tome la columna de las categorias, que es la primera)
    categorias = Reference(pestaña, min_col=min_columna, max_col=min_columna, min_row=min_fila+1, max_row=max_fila)


    #-----ahora solo nos falta crear el gráfico y cargar las referencias dentro del mismo-----
    #creación del gráfico
    barchart = BarChart()
    #agregamos la data (el parámetro titles_from_data = True sirve para indicarle al programa que también debe tomar los headers o encabezados)
    barchart.add_data(data, titles_from_data=True)
    #agregamos las categorias 
    barchart.set_categories(categorias)
    #podemos agregarle un título y un estilo al gráfico
    barchart.title = 'Total de ventas por género'
    barchart.style = 2

    #agregamos el gráfico barchart a la pestaña de excel con la que estamos trabajando, en este caso la posición C10 (el parámetro sirve para indicar en qué celda de la pestaña pondremos el gráfico)
    pestaña.add_chart(barchart, 'C10')


    #----------aplicamos fórmulas de Excel para automatizar los reportes----------

    #si quisiéramos sumar columnas o filas y poner el resultado en una posición específica debemos poner:
    #en la posición B4 se mostrará la suma de la posición B5 hasta la B6
    #pestaña['B7'] = '=SUM(B5:B6)'
    #pestaña['B7'].style = 'Currency' #este estilo hace que se muestre en formato moneda


    #en este caso sumaremos los totales de cada columna de la pestaña de excel que venimos usando, para eso usaremos un bucle for
    #pero antes del for debemos conocer cuáles son las columnas que usaremos, y según vemos en el archivo excel del wb debemos usar de la columna B hasta la G, por eso debemos realizar dos pasos antes del for
    #primero usaremos una variable abecedario para que contenga todos las letras en MAYÚSCULAS (ya que los índices de Excel son en mayúsculas)
    abecedario = list(string.ascii_uppercase)
    #luego nos quedaremos solo con las letras que necesitamos (para eso usamos el slicing o segmentación)
    #en este caso solo debemos usar de la A hasta la G (max_columna)
    abecedario_excel = abecedario[0:max_columna]

    #utilizamos un bucle for para sumar los totales de cada columna
    for i in abecedario_excel :
        # i debe ser distinto de 'A' ya que ahí se encuentran los géneros, y no queremos sumar eso
        if i != 'A' :
            #---sumamos cada total de las columnas especificadas y ponemos el resultado en la posición correspondiente--- 
            #al hacer que se guarde en la columna i y en la fila max_fila+1 nos aseguramos que el programa seguirá funcionando correctamente por más que el excel sea modificado
            #al hacer que se realice la suma entre la columna i y la fila min_fila+1 y la columna i y la fila max_fila nos aseguramos que el programa seguirá realizando las sumas correctamente por más que el excel sea modificado
            pestaña[f'{i}{max_fila+1}'] = f'=SUM({i}{min_fila+1}:{i}{max_fila})'
            pestaña[f'{i}{max_fila+1}'].style = 'Currency' #este estilo hace que se muestre en formato moneda


    #agregamos el nombre 'Total' (para mostrar que en esa fila se mostrarán los totales de todas las sumas anteriores) en la primera columna de la fila max_fila+1 para asegurarnos el correcto funcionamiento sin importar si el excel es modificado
    pestaña[f'{abecedario_excel[0]}{max_fila+1}'] = 'Total'

    #con esto, GENERAMOS UNA AUTOMATIZACIÓN QUE NO AFECTARÁ SU RENDIMIENTO SI EL EXCEL DEL WB ES MODIFICADO
    #de esta forma, en nuestro wb veremos que la última columna se llamará Total y en esa fila estarán todos los totales de las sumas 


    #Por último le damos un poco de formato y estilo al excel al poner un par de títulos o textos descriptivos.
    #-----textos-----
    pestaña['A1'] = 'Reporte automatizado'
    #conseguimos el mes para usarlo (tomamos lo que está antes del punto en mes_extension)
    mes = mes_extension.split('.')[0]
    #hacemos que el mes aparezca en la posicion A2
    pestaña['A2'] = mes
    #-----fuentes de los textos-----
    #de parámetros ponemos el nombre de la fuente, si será negrita o no, y el tamaño
    pestaña['A1'].font = Font('Arial', bold=True, size=20)
    pestaña['A2'].font = Font('Arial', bold=False, size=12)

    #por último, guardamos el archivo wb (el archivo excel creado a partir de la tabla_pivote) 
    wb.save(f'Reporte_{mes_extension}')



#probamos la función con cualquier archivo excel que tenga el formato pedido (en este caso el formato de entrada debe ser Ventas_mes.xlsx)

#automatizar reportes mensuales
automatizar_excel('Ventas_enero.xlsx')
automatizar_excel('Ventas_febrero.xlsx')





